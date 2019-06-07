import openpyxl
import time
import glob
import xml.etree.ElementTree as ET
from openpyxl.styles import Font, Style
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import GREEN
from openpyxl import Workbook
from openpyxl.comments import Comment
from datetime import datetime, date

from collections import Counter

wb = openpyxl.load_workbook('C:\Users\Administrator\My Documents\ExecutionReport.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')


def main():
	
	Listoffiles = glob.glob(r'C:\results\iteration*\output.xml')
	#Listoffiles = glob.glob(r'C:\backup\testresults\output*.xml')
	for filepath in Listoffiles:
		WriteDataFromXMLToXL(filepath)
	
	wb.save('C:\Users\Administrator\My Documents\ExecutionReport.xlsx')
	

def	WriteDataFromXMLToXL(filepath):
	
	tree = ET.parse(filepath)
	root = tree.getroot()
	for suite in root.findall(".//suite/..[@name='Parallel']"):
		tree._setroot(suite)
	
	root = tree.getroot()
	
	TimeList = []
	FailedList = []
	#Parsing the XML and Extracting the data needed in a List of Dictionaries
	for index in range(1,18):
		for child in root:
			if (child.tag == 'suite' and child.attrib['name'] == 'Test Pabot'+str(index) ):
				status = child.find('status')
				starttime = datetime.strptime(status.attrib['starttime'], '%Y%m%d %H:%M:%S.%f' )
				endtime = datetime.strptime(status.attrib['endtime'], '%Y%m%d %H:%M:%S.%f')
				starttime_ts = time.mktime(starttime.timetuple())
				endtime_ts = time.mktime(endtime.timetuple())
				timediffsec = int(endtime_ts - starttime_ts)
				timediff = str(timediffsec/3600) + " hr " + str((timediffsec%3600)/60) + " min " + str(timediffsec%60) + " sec"
				dict1 = {'server': child.attrib['name'][-1],'status': status.attrib['status'], 'timedifference': timediff, 'timediffsec': timediffsec, 'starttime': starttime, 'endtime': endtime}
				TimeList.append(dict1)
				if (dict1['status']=='FAIL'):
					for subelement in child:
						if (subelement.tag == 'test' and subelement.find('status').attrib['status'] == 'FAIL'):
							dict2 = {'server': child.attrib['name'][-1], 'FailedTest': subelement.attrib['name'], 'Failure': subelement.find('status').text}
							FailedList.append(dict2)
							break
	
					
	#outputfileno = filepath[-5]
	outputfileno = filepath.partition('\output.xml')[0][-1]
	print outputfileno
	#Updating the XL with Pass/Fail iteration, Time taken and inserting a comment of start and end time for each Passed iteration 
	for dict in TimeList:
		PassFont = Font(size=12, bold=True, color=GREEN)
		FailFont = Font(size=12, bold=True, color=RED)
		StatusCell = sheet.cell(row=int(outputfileno)+6, column=int(dict['server'])+1)
		TimeCell = sheet.cell(row=int(outputfileno)+111, column=int(dict['server'])+1)
		StatusCell.value = dict['status']
		if (dict['status'] == 'PASS'):
			StatusCell.font = PassFont
			TimeCell.font = PassFont
			TimeCell.value = dict['timedifference']
			commenttext = str(dict['starttime']) + ' - ' + str(dict['endtime'])
			TimeCell.comment = Comment(commenttext,'')
		else:
			StatusCell.font = FailFont
			TimeCell.font = FailFont
			TimeCell.value = 'xx'
			
	#Inserting Comments on those Cells of Excel which failed. Comment is the first reason of Failure
	for dict in FailedList:
		commenttext = 'Test "' + str(dict['FailedTest']) + '" is Failed and failure is \n' +str(dict['Failure'])
		StatusCell = sheet.cell(row=int(outputfileno)+6, column=int(dict['server'])+1)
		if (str(dict['FailedTest']) == 'Status Poll'):
			StatusCell.value = 'No Run'
			StatusCell.font = Font(size=12, bold=True)
		else:
			StatusCell.comment = Comment(commenttext,'')
		

			
	#Updating the XL with Total Time completed for each server in hours mins and secs
	for column in range(2,19,1):
		totalSecCell = sheet.cell(row=215, column=column)
		totalTimeCell = sheet.cell(row=216, column=column)
		totalTimeCell.font = Font(size=13, bold=True)
		for dict in TimeList:
			if (int(dict['server']) == column-1 and dict['status'] == 'PASS'):
				totalvalue = int(totalSecCell.value) + dict['timediffsec']
				totalSecCell.value = totalvalue
				totaltime = str(totalvalue/3600) + " hr " + str((totalvalue%3600)/60) + " min " + str(totalvalue%60) + " sec"
				totalTimeCell.value = totaltime
	
	
main()
